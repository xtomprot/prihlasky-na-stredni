"""
AI Enrichment Workflow for Czech Secondary School Application Data

This script enriches each record in the parsed CSV with additional information:
1. Public transport accessibility from Praha Rajská zahrada (arrival 7:45 on workdays)

Uses web scraping for transport data (idos.cz) and manually curated school addresses/websites.
"""

import argparse
import json
import logging
import re
import time as time_module
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Optional
from urllib.parse import quote

import pandas as pd
import requests

INPUT_CSV = Path(__file__).parent.parent / "output" / "02_parsed_schools.csv"
OUTPUT_CSV = Path(__file__).parent.parent / "output" / "03_enriched_schools.csv"
OUTPUT_XLSX = Path(__file__).parent.parent / "output" / "03_enriched_schools.xlsx"
CONFIG_SCHOOLS = Path(__file__).parent.parent / "config" / "schools_addresses.json"

OUTPUT_CSV.parent.mkdir(parents=True, exist_ok=True)

# Logging setup
LOG_DIR = Path(__file__).parent.parent / "log"
LOG_DIR.mkdir(parents=True, exist_ok=True)
LOG_FILE = LOG_DIR / f"03_enrichment_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"

logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)

# File handler
file_handler = logging.FileHandler(LOG_FILE, encoding="utf-8")
file_handler.setLevel(logging.DEBUG)

# Console handler
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.DEBUG)

# Formatter
formatter = logging.Formatter(
    "%(asctime)s - %(name)s - %(levelname)s - %(message)s", datefmt="%Y-%m-%d %H:%M:%S"
)
file_handler.setFormatter(formatter)
console_handler.setFormatter(formatter)

# Add handlers
logger.addHandler(file_handler)
logger.addHandler(console_handler)

# Configuration (defaults - can be overridden by command-line arguments)
CONFIG = {
    "target_city": "Rajská zahrada",  # Default starting point for public transport
    "target_arrival_time": "07:45",
    "idos_request_delay": 0.5,  # seconds between IDOS requests
    "timeout": 10,  # seconds for HTTP requests
}

# Output columns
OUTPUT_HEADERS = [
    "school_name",
    "curriculum_name",
    "curriculum_code",
    "curriculum_detail",
    "round1_capacity",
    "round1_applications",
    "round1_accepted",
    "round1_acceptance_rate",
    "round1_min_score",
    "round1_avg_score",
    "round2_capacity",
    "round2_applications",
    "round2_accepted",
    "round2_acceptance_rate",
    "round2_min_score",
    "round2_avg_score",
    # New enrichment columns
    "public_transport_available",
    "public_transport_info",
    "public_transport_duration_minutes",
    "public_transport_transfers",
    "school_website",
    "school_phone",
    "school_email",
    "extra_points_grade_avg",
    "extra_points_math",
    "extra_points_extracurricular",
    "min_threshold_czech_language",
    "min_threshold_math",
    "enrollment_notes",
    "street_address",
    # Debug columns
    "debug_idos_url",
    "enrichment_status",
]


class TransportChecker:
    """Check public transport options using IDOS website."""

    def __init__(self, start_point: str = None):
        self.idos_url = "https://idos.cz/pid/spojeni/"
        self.cache = {}
        # Use provided start_point or fall back to config default
        self.target_from = start_point or CONFIG["target_city"]
        self.target_time = CONFIG["target_arrival_time"]  # 07:45
        # IDOS location codes - these may need to be looked up
        self.from_code = "301003"  # Rajská zahrada (estimate)

    def build_idos_url(
        self, destination: str, arrival_code: Optional[str] = None
    ) -> str:
        """
        Build IDOS search URL for journey from Rajská zahrada to school.

        URL format:
        https://idos.cz/pid/spojeni/?date=DD.MM.YYYY&time=07:45&f=FROM&fc=CODE&t=TO&tc=CODE&arr=true&submit=true

        Date is set to next Monday from today.

        Cleans school_name to remove city part (e.g., "School, City, Street" → "School, Street")
        """
        from datetime import datetime, timedelta

        # Handle NaN or None destination
        if pd.isna(destination) or not destination:
            return None

        cleaned_destination = str(destination)

        # Calculate next Monday
        today = datetime.now()
        # weekday(): 0=Monday, 1=Tuesday, ..., 6=Sunday
        days_until_monday = (7 - today.weekday()) % 7
        if days_until_monday == 0:  # If today is Monday, use next Monday
            days_until_monday = 7
        next_monday = today + timedelta(days=days_until_monday)
        date_str = next_monday.strftime("%d.%m.%Y")

        # URL encode the destination
        destination_encoded = quote(cleaned_destination)
        from_encoded = quote(self.target_from)

        url = (
            f"{self.idos_url}"
            f"?date={date_str}"
            f"&time={self.target_time}"
            f"&f={from_encoded}"
            f"&fc={self.from_code}"
            f"&t={destination_encoded}"
            f"&tc={arrival_code or 'null'}"
            f"&arr=true"
            f"&submit=true"
        )

        return url

    def check_transport(
        self, school_name: str, destination: str
    ) -> Optional[Dict[str, Any]]:
        """
        Check public transport from Rajská zahrada to school.

        Extracts from IDOS:
        - Total travel time (from mafra_conn meta tag departure/arrival times)
        - Number of transfers (from HTML "Přesun" elements)

        Returns dict with:
        - available: bool (True if route found)
        - duration_minutes: int (total travel time)
        - transfers: int (number of transfer points)
        - route_info: str (summary)
        - idos_url: str (link to full results)
        - constructed_address: str (cleaned school address used for IDOS query)
        """
        import base64
        from datetime import timedelta

        cache_key = f"{school_name}_{destination}"
        if cache_key in self.cache:
            logger.debug(f"Using cached transport info for {school_name}")
            return self.cache[cache_key]

        try:
            cleaned_destination = destination or school_name

            # Build the IDOS search URL
            idos_url = self.build_idos_url(cleaned_destination)
            logger.debug(f"IDOS lookup: {self.target_from} → {cleaned_destination}")

            # Fetch IDOS page
            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
            }

            response = requests.get(
                idos_url, headers=headers, timeout=CONFIG["timeout"]
            )
            response.raise_for_status()

            duration_minutes = None
            num_transfers = None
            route_summary = None

            # ===== STRATEGY 1: Extract from mafra_conn meta tag =====
            # This contains departure and arrival times for the first journey
            mafra_match = re.search(
                r'name="mafra_conn"\s+content="([^"]+)"', response.text
            )
            if mafra_match:
                try:
                    encoded = mafra_match.group(1)
                    decoded = base64.b64decode(encoded).decode("utf-8")
                    conn_data = json.loads(decoded)

                    # Extract and parse departure/arrival times
                    cas_odj = conn_data.get("cas_odj")  # e.g., "6:31"
                    cas_pri = conn_data.get("cas_pri")  # e.g., "7:29"

                    if cas_odj and cas_pri:
                        # Calculate duration in minutes
                        from datetime import datetime as dt

                        time_format = "%H:%M"
                        dep_time = dt.strptime(cas_odj, time_format)
                        arr_time = dt.strptime(cas_pri, time_format)

                        # Handle next-day arrival (shouldn't happen with same date)
                        if arr_time < dep_time:
                            from datetime import timedelta

                            arr_time += timedelta(days=1)

                        duration_sec = (arr_time - dep_time).total_seconds()
                        duration_minutes = int(duration_sec / 60)

                        logger.debug(
                            f"Extracted duration from mafra_conn: {duration_minutes} min ({cas_odj} → {cas_pri})"
                        )

                except Exception as e:
                    logger.warning(f"Could not extract time from mafra_conn: {e}")

            # ===== STRATEGY 2: Extract number of transfers from HTML =====
            # Count "Přesun" elements which indicate transfer points
            presun_matches = re.findall(
                r"Přesun\s+asi\s+(\d+)\s*(?:minut|min)", response.text, re.IGNORECASE
            )

            if presun_matches:
                # Each "Přesun" entry indicates a transfer point
                num_transfers = len(presun_matches)
                total_transfer_time = sum(int(m) for m in presun_matches)
                logger.debug(
                    f"Found {num_transfers} transfer(s) with {total_transfer_time} min total connection time"
                )
            else:
                # Check for "bez přestupu" (no transfers/direct route)
                if "bez přestupu" in response.text.lower():
                    num_transfers = 0
                    logger.debug("Found 'bez přestupu' - direct route (0 transfers)")

            # Build result
            result = {
                "available": duration_minutes
                is not None,  # True if we got duration data
                "duration_minutes": duration_minutes,
                "transfers": num_transfers,
                "idos_url": idos_url,
                "constructed_address": cleaned_destination,
                "data_source": "idos_mafra_conn",
            }

            if duration_minutes is not None:
                if num_transfers is not None:
                    route_summary = f"{duration_minutes} min, {num_transfers} transfer{'s' if num_transfers != 1 else ''}"
                else:
                    route_summary = f"{duration_minutes} min"
                result["route_info"] = route_summary
            else:
                result["route_info"] = "Could not extract timing info"

            logger.info(
                f"Transport check for {school_name}: {route_summary or 'No data'}"
            )

            self.cache[cache_key] = result
            return result

        except requests.exceptions.RequestException as e:
            logger.warning(f"IDOS lookup failed for {school_name}: {e}")

            cleaned_destination = destination or school_name
            result = {
                "available": None,
                "duration_minutes": None,
                "transfers": None,
                "route_info": "Lookup failed",
                "idos_url": self.build_idos_url(cleaned_destination),
                "constructed_address": cleaned_destination,
                "data_source": "lookup_failed",
            }

            self.cache[cache_key] = result
            return result

        except Exception as e:
            logger.exception(f"Error checking transport for {school_name}: {e}")
            cleaned_destination = destination or school_name
            return {
                "available": None,
                "duration_minutes": None,
                "transfers": None,
                "route_info": None,
                "idos_url": None,
                "constructed_address": cleaned_destination,
                "data_source": "error",
            }


def load_schools_config(config_file: Path) -> Dict[str, Dict[str, str]]:
    """
    Load school configuration from JSON file.

    Returns dict with school_name as key and {website, address} as value.
    """
    if not config_file.exists():
        logger.warning(f"Config file not found: {config_file}")
        return {}

    try:
        with open(config_file, "r", encoding="utf-8") as f:
            config = json.load(f)
        logger.info(f"Loaded config for {len(config)} schools from {config_file.name}")
        return config
    except Exception as e:
        logger.error(f"Error loading config file: {e}", exc_info=True)
        return {}


def enrich_record(
    row: pd.Series,
    transport_checker: TransportChecker,
    schools_config: Dict[str, Dict[str, str]],
) -> Dict[str, Any]:
    """
    Enrich a single record with transport data using school configuration.

    Uses school_name as foreign key to look up website and address from config.
    Returns a dictionary with enrichment fields.
    """

    enrichment = {
        "public_transport_available": None,
        "public_transport_info": None,
        "public_transport_duration_minutes": None,
        "public_transport_transfers": None,
        "school_website": None,
        "school_phone": None,
        "school_email": None,
        "extra_points_grade_avg": None,
        "extra_points_math": None,
        "extra_points_extracurricular": None,
        "min_threshold_czech_language": None,
        "min_threshold_math": None,
        "enrollment_notes": None,
        "street_address": None,
        "debug_idos_url": None,
        "enrichment_status": "pending",
    }

    school_name = row["school_name"]

    # Handle NaN values
    if pd.isna(school_name):
        school_name = "Unknown School"
    else:
        school_name = str(school_name)

    try:
        # Look up school data using school_name as foreign key
        school_data = schools_config.get(school_name, {})
        school_website = school_data.get("website")
        street_address = school_data.get("address")

        enrichment["school_website"] = school_website
        enrichment["street_address"] = street_address

        # Check public transport using address from config
        destination = street_address or school_name
        logger.debug(
            f"Checking transport for {school_name} (destination: {destination})..."
        )
        transport = transport_checker.check_transport(school_name, destination)
        if transport:
            enrichment["public_transport_available"] = transport.get("available", False)
            enrichment["public_transport_info"] = transport.get("route_info")
            enrichment["public_transport_duration_minutes"] = transport.get(
                "duration_minutes"
            )
            enrichment["public_transport_transfers"] = transport.get("transfers")
            enrichment["debug_idos_url"] = transport.get("idos_url")

        time_module.sleep(CONFIG["idos_request_delay"])

        enrichment["enrichment_status"] = "completed"

    except Exception as e:
        logger.error(f"Error enriching {school_name}: {e}", exc_info=True)
        enrichment["enrichment_status"] = f"error: {str(e)}"

    return enrichment


def main():
    """Main enrichment workflow."""

    # Parse command-line arguments
    parser = argparse.ArgumentParser(
        description="Enrich Czech secondary school data with public transport information"
    )
    parser.add_argument(
        "--pid-stop",
        type=str,
        default="Rajská zahrada",
        help="Starting point for public transport route queries (default: Rajská zahrada)",
    )
    args = parser.parse_args()

    # Update CONFIG with the specified starting point
    if args.pid_stop != "Rajská zahrada":
        CONFIG["target_city"] = args.pid_stop
        logger.info(f"Using custom PID starting point: {args.pid_stop}")
    else:
        logger.info("Using default PID starting point: Rajská zahrada")

    logger.info("=" * 80)
    logger.info("Starting AI Enrichment Workflow")
    logger.info("=" * 80)

    if not INPUT_CSV.exists():
        logger.error(f"Input file not found: {INPUT_CSV}")
        return

    logger.info(f"Reading: {INPUT_CSV}")

    try:
        df = pd.read_csv(INPUT_CSV, encoding="utf-8")
    except Exception as e:
        logger.error(f"Error reading CSV: {e}", exc_info=True)
        return

    logger.info(f"Loaded {len(df)} records")

    # Get unique schools by school_name
    unique_schools = df.drop_duplicates(subset=["school_name"]).copy()
    logger.info(
        f"Found {len(unique_schools)} unique schools (out of {len(df)} total programme rows)"
    )

    # Initialize enrichment tools
    logger.debug("Initializing enrichment tools...")
    transport_checker = TransportChecker(start_point=CONFIG["target_city"])
    schools_config = load_schools_config(CONFIG_SCHOOLS)

    # Enrich each unique school once
    school_enrichments = {}  # Cache enrichments by school_name

    for idx, (_, school_row) in enumerate(unique_schools.iterrows(), 1):
        school_name = school_row["school_name"]
        logger.info(f"[{idx}/{len(unique_schools)}] Enriching school: {school_name}")

        # Get enrichment data once per unique school
        enrichment = enrich_record(school_row, transport_checker, schools_config)

        # Cache the enrichment for this school
        school_enrichments[school_name] = enrichment

    # Now join enrichments back to all programmes for each school
    enriched_records = []
    for idx, (_, row) in enumerate(df.iterrows(), 1):
        school_name = row["school_name"]

        # Get the cached enrichment for this school
        enrichment = school_enrichments.get(school_name, {})

        # Combine original and enrichment data
        enriched_row = row.to_dict()
        enriched_row.update(enrichment)
        enriched_records.append(enriched_row)

    # Create enriched dataframe
    enriched_df = pd.DataFrame(enriched_records)

    # Ensure all columns are present
    for col in OUTPUT_HEADERS:
        if col not in enriched_df.columns:
            enriched_df[col] = None

    # Reorder columns
    enriched_df = enriched_df[OUTPUT_HEADERS]

    # Write enriched CSV
    try:
        logger.debug(f"Writing enriched CSV to {OUTPUT_CSV}")
        enriched_df.to_csv(OUTPUT_CSV, index=False, encoding="utf-8")
        logger.info(f"Written {len(enriched_df)} enriched records to {OUTPUT_CSV}")
    except Exception as e:
        logger.error(f"Error writing CSV: {e}", exc_info=True)

    # Write final XLSX copy (preserves Czech diacritics)
    try:
        logger.debug(f"Writing enriched XLSX to {OUTPUT_XLSX}")
        enriched_df.to_excel(OUTPUT_XLSX, index=False, engine="openpyxl")

        # Add hyperlinks to URL columns
        from openpyxl import load_workbook
        from openpyxl.styles import Font

        wb = load_workbook(OUTPUT_XLSX)
        ws = wb.active

        # Find URL column indices (1-based for Excel)
        url_columns = {
            "school_website": None,
            "debug_idos_url": None,
        }
        for col_idx, col_name in enumerate(OUTPUT_HEADERS, 1):
            if col_name in url_columns:
                url_columns[col_name] = col_idx

        # Add hyperlinks starting from row 2 (row 1 is header)
        for col_name, col_idx in url_columns.items():
            if not col_idx:
                continue
            for row_idx in range(2, len(enriched_df) + 2):
                cell = ws.cell(row=row_idx, column=col_idx)
                url = cell.value

                # Only add hyperlink if cell contains a URL
                if (
                    url
                    and isinstance(url, str)
                    and (url.startswith("http://") or url.startswith("https://"))
                ):
                    cell.hyperlink = url
                    cell.font = Font(color="0563C1", underline="single")
                    logger.debug(f"Added hyperlink to {url}")

        wb.save(OUTPUT_XLSX)
        logger.info(f"Written {len(enriched_df)} enriched records to {OUTPUT_XLSX}")
        logger.debug("School website URLs formatted as clickable hyperlinks")
    except Exception as e:
        logger.error(f"Error writing XLSX: {e}", exc_info=True)

    logger.info("=" * 80)
    logger.info("Enrichment Complete")
    logger.info(f"  Total programme rows: {len(enriched_records)}")
    logger.info(f"  Unique schools enriched: {len(school_enrichments)}")
    logger.info("=" * 80)


if __name__ == "__main__":
    main()
